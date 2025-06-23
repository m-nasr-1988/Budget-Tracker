# Streamlit Budget Tracker (Basic Editable Table Version)
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

st.set_page_config(page_title="Budget Tracker", layout="wide")

# Constants
DATA_FILE = Path("data/budget_tracker.xlsx")
TEMPLATE_SHEET_PREFIX = "Template_"
COLUMNS = ["Type", "Category", "Description", "Amount"]
TYPE_OPTIONS = ["Income", "Expense"]
CATEGORY_OPTIONS = ["Salary", "Rent", "Groceries", "Utilities", "Fuel", "Phone", "Internet", "Insurance", "Gym", "Camp", "Other"]

# Ensure Excel exists
def init_excel():
    if not DATA_FILE.exists():
        with pd.ExcelWriter(DATA_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=COLUMNS).to_excel(writer, sheet_name="Sheet1", index=False)

def load_excel():
    return pd.ExcelFile(DATA_FILE)

def read_sheet(sheet_name):
    xl = load_excel()
    if sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = df[COLUMNS].fillna("")
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
        return df
    return pd.DataFrame(columns=COLUMNS)

def save_to_excel(sheet_name, df):
    with pd.ExcelWriter(DATA_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Init file
init_excel()

# Sidebar UI
st.sidebar.title("📊 Budget Tracker")
existing_sheets = load_excel().sheet_names if DATA_FILE.exists() else []
months = [s for s in existing_sheets if not s.startswith(TEMPLATE_SHEET_PREFIX)]
templates = [s.replace(TEMPLATE_SHEET_PREFIX, "") for s in existing_sheets if s.startswith(TEMPLATE_SHEET_PREFIX)]

selected_month = st.sidebar.selectbox("Select Month", ["+ New Month"] + months)

# Delete sheet checkbox
if selected_month != "+ New Month" and st.sidebar.checkbox("🗑 Delete This Month's Data"):
    wb = load_workbook(DATA_FILE)
    if selected_month in wb.sheetnames:
        del wb[selected_month]
        wb.save(DATA_FILE)
        st.success(f"Deleted sheet: {selected_month}")
        st.experimental_rerun()

# Create new month
if selected_month == "+ New Month":
    new_month_name = st.sidebar.text_input("Enter new month name (e.g., July - 2025)")
    use_template = st.sidebar.checkbox("Use a template?")
    selected_template = st.sidebar.selectbox("Select template", ["None"] + templates) if use_template else None

    if st.sidebar.button("➕ Create Month") and new_month_name:
        if use_template and selected_template != "None":
            df_month = read_sheet(TEMPLATE_SHEET_PREFIX + selected_template)
            df_month["Amount"] = 0.0
        else:
            df_month = pd.DataFrame([
                {"Type": "Income", "Category": "", "Description": "", "Amount": 0.0}
            ])
        save_to_excel(new_month_name, df_month)

        # Remove "Sheet1" if it exists after first month is created
        wb = load_workbook(DATA_FILE)
        if "Sheet1" in wb.sheetnames:
            del wb["Sheet1"]
            wb.save(DATA_FILE)

        st.session_state["selected_month"] = new_month_name
        st.rerun()
else:
    st.session_state["selected_month"] = selected_month

# Active sheet
current_month = st.session_state.get("selected_month")
if current_month and current_month != "+ New Month":
    st.header(f"📅 Editing: {current_month}")

    if "df" not in st.session_state:
        st.session_state.df = read_sheet(current_month)
        if st.session_state.df.empty:
            st.session_state.df = pd.DataFrame([
                {"Type": "Income", "Category": "Salary", "Description": "", "Amount": 0.0}
            ])

    # Work on a copy for editing (display only)
    df_display = st.session_state.df.copy()

    # Add a Row number for display only
    df_display.insert(0, "Row", range(1, len(df_display) + 1))

    # Show editable table (excluding the Row column from being editable)
    edited_df = st.data_editor(
        df_display,
        num_rows="dynamic",
        use_container_width=True,
        key="data_editor",
        column_config={
            "Row": st.column_config.NumberColumn("Row", disabled=True),
            "Type": st.column_config.SelectboxColumn("Type", options=TYPE_OPTIONS),
            "Category": st.column_config.SelectboxColumn("Category", options=CATEGORY_OPTIONS),
            "Description": st.column_config.TextColumn("Description"),
            "Amount": st.column_config.NumberColumn("Amount", format="%.2f"),
        }
    )

    # Remove the Row column before saving back
    edited_df = edited_df.drop(columns=["Row"])

    # Clean up Amount field
    edited_df["Amount"] = pd.to_numeric(edited_df["Amount"], errors="coerce").fillna(0.0)

    st.session_state.df = edited_df.reset_index(drop=True)

    cleaned_df = st.session_state.df.dropna(how="all")
    cleaned_df["Amount"] = pd.to_numeric(cleaned_df["Amount"], errors="coerce").fillna(0.0)

    if st.button("💾 Save to Excel"):
        cleaned_df = st.session_state.df.dropna(how="all")
        save_to_excel(current_month, cleaned_df)
        st.success(f"Saved to {current_month} ✅")

    with st.expander("🗂 Save as Template"):
        template_name = st.text_input("Template name")
        if st.button("📁 Save Template") and template_name:
            df_template = st.session_state.df.copy()
            df_template["Amount"] = 0.0
            save_to_excel(TEMPLATE_SHEET_PREFIX + template_name, df_template)
            st.success(f"Saved template: {template_name}")

    col1, col2 = st.columns(2)
    with col1:
        total_in = st.session_state.df[st.session_state.df["Type"] == "Income"]["Amount"].sum()
        total_out = st.session_state.df[st.session_state.df["Type"] == "Expense"]["Amount"].sum()
        st.metric("💰 Total In", f"{total_in:,.0f}")
        st.metric("💸 Total Out", f"{total_out:,.0f}")
    with col2:
        st.metric("📉 Remaining", f"{(total_in - total_out):,.0f}")

    st.subheader("📊 Breakdown by Category")
    if not st.session_state.df.empty:
        category_summary = st.session_state.df[st.session_state.df["Type"] == "Expense"].groupby("Category")["Amount"].sum().sort_values(ascending=False)
        st.bar_chart(category_summary)
else:
    st.info("👈 Please select or create a month from the sidebar.")
